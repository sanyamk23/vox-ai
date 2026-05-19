"""
Campaign management views — bulk Excel upload, sequential calling, live stats, Excel export.
"""
from __future__ import annotations

import asyncio
import io
import json
import logging
import os
import re

from asgiref.sync import sync_to_async
from django.core.cache import cache

# sync_to_async wrapper safe for background asyncio tasks that outlive the HTTP
# request that created them.  thread_sensitive=False uses the global thread pool
# instead of the request-bound CurrentThreadExecutor (which is torn down when the
# view returns, causing "CurrentThreadExecutor already quit or is broken" errors).
def _bg(fn):
    return sync_to_async(fn, thread_sensitive=False)
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

logger = logging.getLogger(__name__)

_E164_RE = re.compile(r"^\+[1-9]\d{6,14}$")

# Strong-reference set keeps background tasks alive (not GC-cancelled)
_CAMPAIGN_TASKS: dict[int, asyncio.Task] = {}


# ---------------------------------------------------------------------------
# Phone helpers
# ---------------------------------------------------------------------------

def _normalize_phone(raw: str) -> str:
    """Best-effort normalisation to E.164. Returns raw if nothing matches."""
    phone = re.sub(r"[\s\-\(\)\.\+]", "", str(raw or "").strip())
    if not phone:
        return raw
    # Already has country prefix (10+ digits)
    if re.match(r"^\+?91[6-9]\d{9}$", phone):
        return f"+91{phone[-10:]}"
    if re.match(r"^[6-9]\d{9}$", phone):          # Indian mobile, no prefix
        return f"+91{phone}"
    if re.match(r"^\d{10}$", phone):               # US 10-digit
        return f"+1{phone}"
    if re.match(r"^\+?1\d{10}$", phone):           # US with country code
        return f"+1{phone[-10:]}"
    # Already well-formed E.164 (starts with + in original)
    if raw.strip().startswith("+") and re.match(r"^\d{7,14}$", phone):
        return f"+{phone}"
    return raw                                      # give up; will fail E.164 check


def _validate_candidates(rows: list[dict]) -> dict:
    valid, invalid, duplicates = [], [], []
    seen: set[str] = set()

    for row in rows:
        name = str(row.get("name") or "").strip()
        raw_phone = str(row.get("phone") or "").strip()

        if not name and not raw_phone:
            continue                                # blank row

        phone = _normalize_phone(raw_phone)
        error = None

        if not name:
            error = "Missing name"
        elif not raw_phone:
            error = "Missing phone number"
        elif not _E164_RE.match(phone):
            error = f"Invalid phone: {raw_phone}"

        entry = {"name": name, "phone": phone, "raw_phone": raw_phone}

        if error:
            invalid.append({**entry, "error": error})
        elif phone in seen:
            duplicates.append({**entry, "error": "Duplicate phone number"})
        else:
            seen.add(phone)
            valid.append(entry)

    return {"valid": valid, "invalid": invalid, "duplicates": duplicates}


# ---------------------------------------------------------------------------
# CRUD endpoints
# ---------------------------------------------------------------------------

@csrf_exempt
def campaign_list_or_create(request):
    """GET /api/campaigns/ → list; POST /api/campaigns/ → create (alias for /create/)."""
    if request.method == 'POST':
        return create_campaign(request)
    return list_campaigns(request)


@csrf_exempt
@require_http_methods(["GET"])
def list_campaigns(request):
    """GET /api/campaigns/"""
    from .models import Campaign
    campaigns = []
    for c in Campaign.objects.all():
        s = c.stats
        campaigns.append({
            "id": c.id,
            "name": c.name,
            "status": c.status,
            "voice_id": c.voice_id,
            "total_uploaded": c.total_uploaded,
            "valid_count": c.valid_count,
            "stats": s,
            "created_at": c.created_at.isoformat(),
            "started_at": c.started_at.isoformat() if c.started_at else None,
        })
    return JsonResponse({"campaigns": campaigns})


@csrf_exempt
@require_http_methods(["POST"])
def create_campaign(request):
    """
    POST /api/campaigns/
    Multipart form: file (xlsx), campaign_name, job_description,
                    voice_id, delay_seconds
    """
    try:
        import openpyxl
    except ImportError:
        return JsonResponse({"error": "openpyxl not installed on server"}, status=500)

    from .models import Campaign, CampaignCandidate

    excel_file = request.FILES.get("file")
    if not excel_file:
        return JsonResponse({"error": "No Excel file uploaded"}, status=400)

    campaign_name = (request.POST.get("campaign_name") or "").strip() or f"Campaign {timezone.now():%Y-%m-%d %H:%M}"
    jd            = (request.POST.get("job_description") or "").strip()
    voice_id      = (request.POST.get("voice_id") or "priya").strip()
    delay_seconds = _safe_int(request.POST.get("delay_seconds"), 30)

    # Parse workbook
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        return JsonResponse({"error": f"Cannot read Excel file: {exc}"}, status=400)

    if not all_rows:
        return JsonResponse({"error": "Excel file is empty"}, status=400)

    # Detect Name / Phone columns from header row
    header = [str(h or "").strip().lower() for h in all_rows[0]]
    name_col  = next((i for i, h in enumerate(header) if "name" in h), None)
    phone_col = next(
        (i for i, h in enumerate(header) if any(k in h for k in ("phone", "mobile", "number", "contact", "tel"))),
        None,
    )

    if name_col is None or phone_col is None:
        return JsonResponse(
            {"error": "Could not find 'Name' and 'Phone' columns. Please ensure your Excel has these headers."},
            status=400,
        )

    # Build row dicts
    data_rows = []
    for row in all_rows[1:]:
        if not row:
            continue
        n_val = row[name_col]  if len(row) > name_col  else None
        p_val = row[phone_col] if len(row) > phone_col else None
        data_rows.append({"name": n_val, "phone": p_val})

    result = _validate_candidates(data_rows)
    total  = len(data_rows)

    # Persist campaign
    campaign = Campaign.objects.create(
        name=campaign_name,
        job_description=jd,
        voice_id=voice_id,
        delay_seconds=delay_seconds,
        total_uploaded=total,
        valid_count=len(result["valid"]),
        invalid_count=len(result["invalid"]),
        duplicate_count=len(result["duplicates"]),
    )

    # Bulk-create candidates (valid + invalid for audit)
    to_create = [
        CampaignCandidate(
            campaign=campaign, name=c["name"], phone=c["phone"],
            is_valid=True, is_duplicate=False,
        )
        for c in result["valid"]
    ] + [
        CampaignCandidate(
            campaign=campaign, name=c.get("name", ""), phone=c.get("raw_phone", c.get("phone", "")),
            is_valid=False, is_duplicate=False, validation_error=c.get("error", ""),
        )
        for c in result["invalid"]
    ] + [
        CampaignCandidate(
            campaign=campaign, name=c.get("name", ""), phone=c.get("raw_phone", c.get("phone", "")),
            is_valid=False, is_duplicate=True, validation_error="Duplicate",
        )
        for c in result["duplicates"]
    ]
    CampaignCandidate.objects.bulk_create(to_create)

    return JsonResponse({
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "validation": {
            "total_uploaded": total,
            "valid": len(result["valid"]),
            "invalid": len(result["invalid"]),
            "duplicates": len(result["duplicates"]),
            "invalid_details": [
                {"name": c.get("name"), "phone": c.get("raw_phone", ""), "error": c.get("error")}
                for c in result["invalid"]
            ],
            "duplicate_details": [
                {"name": c.get("name"), "phone": c.get("raw_phone", "")}
                for c in result["duplicates"]
            ],
        },
        "candidates": [
            {"name": c["name"], "phone": c["phone"]}
            for c in result["valid"]
        ],
    }, status=201)


@require_http_methods(["GET"])
def get_campaign(request, campaign_id):
    """GET /api/campaigns/<id>/"""
    from .models import Campaign
    try:
        campaign = Campaign.objects.get(id=campaign_id)
    except Campaign.DoesNotExist:
        return JsonResponse({"error": "Campaign not found"}, status=404)

    candidates = list(
        campaign.candidates
        .filter(is_valid=True, is_duplicate=False)
        .values("id", "name", "phone", "status", "call_outcome",
                "interest_level", "call_duration", "ai_summary", "called_at", "ended_at")
    )
    # Serialise datetimes
    for c in candidates:
        c["called_at"] = c["called_at"].isoformat() if c["called_at"] else None
        c["ended_at"]  = c["ended_at"].isoformat()  if c["ended_at"]  else None

    return JsonResponse({
        "id": campaign.id,
        "name": campaign.name,
        "status": campaign.status,
        "voice_id": campaign.voice_id,
        "delay_seconds": campaign.delay_seconds,
        "job_description": campaign.job_description,
        "validation_summary": {
            "total_uploaded": campaign.total_uploaded,
            "valid": campaign.valid_count,
            "invalid": campaign.invalid_count,
            "duplicates": campaign.duplicate_count,
        },
        "stats": campaign.stats,
        "candidates": candidates,
        "created_at": campaign.created_at.isoformat(),
        "started_at": campaign.started_at.isoformat() if campaign.started_at else None,
        "completed_at": campaign.completed_at.isoformat() if campaign.completed_at else None,
    })


@csrf_exempt
@require_http_methods(["POST"])
async def start_campaign(request, campaign_id):
    """POST /api/campaigns/<id>/start/"""
    from .models import Campaign
    try:
        campaign = await sync_to_async(Campaign.objects.get)(id=campaign_id)
    except Campaign.DoesNotExist:
        return JsonResponse({"error": "Campaign not found"}, status=404)

    if campaign.status == Campaign.RUNNING:
        # Allow re-start if the task died (e.g. container restart wiped _CAMPAIGN_TASKS)
        existing = _CAMPAIGN_TASKS.get(campaign_id)
        if existing and not existing.done():
            return JsonResponse({"error": "Campaign is already running"}, status=400)
        # Task is gone — fall through and re-create it
    if campaign.valid_count == 0:
        return JsonResponse({"error": "No valid candidates to call"}, status=400)

    campaign.status     = Campaign.RUNNING
    campaign.started_at = campaign.started_at or timezone.now()
    await sync_to_async(campaign.save)(update_fields=["status", "started_at"])

    # Reset any candidates stuck in CALLING state from a previous interrupted run
    from .models import CampaignCandidate
    reset_count = await sync_to_async(
        CampaignCandidate.objects.filter(
            campaign_id=campaign_id,
            status=CampaignCandidate.CALLING,
        ).update
    )(status=CampaignCandidate.PENDING)
    if reset_count:
        logger.info("[Campaign-%d] Reset %d stuck CALLING candidate(s) to PENDING", campaign_id, reset_count)

    # Launch background caller (idempotent — cancel old task first)
    old = _CAMPAIGN_TASKS.get(campaign_id)
    if old and not old.done():
        old.cancel()

    task = asyncio.create_task(_run_campaign_caller(campaign_id))
    _CAMPAIGN_TASKS[campaign_id] = task

    return JsonResponse({"status": "started", "campaign_id": campaign_id})


@csrf_exempt
@require_http_methods(["POST"])
def pause_campaign(request, campaign_id):
    """POST /api/campaigns/<id>/pause/"""
    from .models import Campaign
    try:
        campaign = Campaign.objects.get(id=campaign_id)
    except Campaign.DoesNotExist:
        return JsonResponse({"error": "Campaign not found"}, status=404)

    campaign.status = Campaign.PAUSED
    campaign.save(update_fields=["status"])

    task = _CAMPAIGN_TASKS.get(campaign_id)
    if task and not task.done():
        task.cancel()

    return JsonResponse({"status": "paused"})


@require_http_methods(["GET"])
def export_campaign(request, campaign_id):
    """GET /api/campaigns/<id>/export/ — download Excel report."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return JsonResponse({"error": "openpyxl not installed"}, status=500)

    from .models import Campaign
    try:
        campaign = Campaign.objects.get(id=campaign_id)
    except Campaign.DoesNotExist:
        return JsonResponse({"error": "Campaign not found"}, status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screening Report"

    headers = [
        "#", "Name", "Phone", "Status", "Call Outcome", "Interest",
        "Duration (min)", "AI Summary", "Called At", "Ended At",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1a1a2e")
        cell.alignment = Alignment(horizontal="center")

    STATUS_COLOURS = {
        "completed": "D4EDDA",
        "failed":    "F8D7DA",
        "calling":   "FFF3CD",
        "pending":   "E2E3E5",
    }

    for i, c in enumerate(
        campaign.candidates.filter(is_valid=True, is_duplicate=False).order_by("created_at"), 1
    ):
        duration = round(c.call_duration / 60, 1) if c.call_duration else ""
        row_vals = [
            i, c.name, c.phone, c.status.upper(), c.call_outcome or "",
            c.interest_level or "", duration, c.ai_summary or "",
            c.called_at.strftime("%Y-%m-%d %H:%M") if c.called_at else "",
            c.ended_at.strftime("%Y-%m-%d %H:%M")  if c.ended_at  else "",
        ]
        ws.append(row_vals)
        colour = STATUS_COLOURS.get(c.status, "FFFFFF")
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor=colour)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in col) + 4, 60
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"campaign_{campaign.name.replace(' ', '_')}_{timezone.now():%Y%m%d}.xlsx"
    resp  = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


# ---------------------------------------------------------------------------
# Background caller
# ---------------------------------------------------------------------------

async def _run_campaign_caller(campaign_id: int) -> None:
    """Sequential async caller — one candidate at a time with configurable delay."""
    from .models import Campaign, CampaignCandidate

    logger.info("[Campaign-%d] Caller task started", campaign_id)

    try:
        host_url    = os.getenv("PUBLIC_URL", "").strip()
        from_number = os.getenv("TWILIO_PHONE_NUMBER", "").strip()

        if not host_url or not from_number:
            logger.error("[Campaign-%d] PUBLIC_URL or TWILIO_PHONE_NUMBER not configured", campaign_id)
            await _bg(Campaign.objects.filter(id=campaign_id).update)(status=Campaign.PAUSED)
            return

        cid = campaign_id  # stable local for lambdas in this scope
        while True:
            campaign = await _bg(Campaign.objects.get)(id=cid)
            if campaign.status != Campaign.RUNNING:
                logger.info("[Campaign-%d] Status=%s — stopping", cid, campaign.status)
                break

            # Next pending candidate — explicit order so we always call in upload order
            candidate = await _bg(
                lambda: CampaignCandidate.objects.filter(
                    campaign_id=cid,
                    is_valid=True,
                    is_duplicate=False,
                    status=CampaignCandidate.PENDING,
                ).order_by("id").first()
            )()

            if candidate is None:
                logger.info("[Campaign-%d] All candidates processed — marking completed", cid)
                await _bg(Campaign.objects.filter(id=cid).update)(
                    status=Campaign.COMPLETED,
                    completed_at=timezone.now(),
                )
                break

            logger.info("[Campaign-%d] Next candidate: %s (%s)", cid, candidate.name, candidate.phone)

            # Mark as calling
            candidate.status    = CampaignCandidate.CALLING
            candidate.called_at = timezone.now()
            await _bg(candidate.save)(update_fields=["status", "called_at"])

            call_sid = ""
            try:
                from .views import _place_call
                result = await _bg(_place_call)(
                    to_number=candidate.phone,
                    from_number=from_number,
                    host_url=host_url,
                    jd=campaign.job_description or "Screening call",
                    name=candidate.name,
                    voice_id=campaign.voice_id,
                )
                call_sid = result.get("call_sid") or ""
                if not call_sid:
                    raise RuntimeError("_place_call returned no call_sid")

                candidate.call_sid = call_sid
                await _bg(candidate.save)(update_fields=["call_sid"])

                # Store link so the status webhook can update the candidate
                await _bg(cache.set)(
                    f"vox:campaign_call:{call_sid}",
                    {"campaign_id": cid, "candidate_id": candidate.id},
                    timeout=3600,
                )
                logger.info("[Campaign-%d] Placed call to %s → SID %s", cid, candidate.name, call_sid)

                # Wait for call to finish before dialling the next candidate
                await _wait_for_call_end(call_sid, timeout=600)
                logger.info("[Campaign-%d] Call ended for %s (SID %s)", cid, candidate.name, call_sid)
                await _sync_call_result(candidate.id, call_sid)

            except asyncio.CancelledError:
                raise
            except Exception as exc:
                logger.error("[Campaign-%d] Call failed for %s: %s", cid, candidate.name, exc, exc_info=True)
                candidate.status = CampaignCandidate.FAILED
                await _bg(candidate.save)(update_fields=["status"])

            # Brief pause between calls
            delay = campaign.delay_seconds
            logger.info("[Campaign-%d] Sleeping %ds before next candidate", cid, delay)
            await asyncio.sleep(delay)

    except asyncio.CancelledError:
        logger.info("[Campaign-%d] Caller task cancelled", campaign_id)
    except Exception as exc:
        logger.error("[Campaign-%d] Caller crashed: %s", campaign_id, exc, exc_info=True)


async def _wait_for_call_end(call_sid: str, timeout: int = 480) -> None:
    """Poll until the call ends, using two independent signals."""
    from .models import CallSession, CampaignCandidate
    sid = call_sid

    def _check_ended():
        ended = (
            CallSession.objects.filter(call_sid=sid)
                               .values_list("ended_at", flat=True)
                               .first()
        )
        if ended is not None:
            return True
        # Secondary: webhook already moved candidate out of CALLING
        status = (
            CampaignCandidate.objects.filter(call_sid=sid)
                                     .values_list("status", flat=True)
                                     .first()
        )
        return bool(status and status != CampaignCandidate.CALLING)

    elapsed = 0
    poll = 3  # check every 3 seconds — faster response
    while elapsed < timeout:
        await asyncio.sleep(poll)
        elapsed += poll
        if await _bg(_check_ended)():
            logger.info("[Campaign] Call %s ended after ~%ds", sid, elapsed)
            return
    logger.warning("[Campaign] Call %s timed out after %ds — moving on", sid, timeout)


async def _sync_call_result(candidate_id: int, call_sid: str) -> None:
    from .models import CallSession, CampaignCandidate
    sid = call_sid
    try:
        session = await _bg(lambda: CallSession.objects.filter(call_sid=sid).first())()
        candidate = await _bg(CampaignCandidate.objects.get)(id=candidate_id)

        if not session:
            candidate.status = CampaignCandidate.FAILED
            await _bg(candidate.save)(update_fields=["status"])
            return

        if session.created_at and session.ended_at:
            candidate.call_duration = int((session.ended_at - session.created_at).total_seconds())

        candidate.call_outcome   = session.call_outcome or ""
        candidate.interest_level = session.call_outcome or ""
        candidate.transcript     = session.transcript or []
        candidate.notes          = session.notes or {}
        candidate.ended_at       = session.ended_at

        if session.notes:
            bullets = session.notes.get("summary_bullets", [])
            candidate.ai_summary = " | ".join(bullets[:3]) if bullets else ""

        candidate.status = CampaignCandidate.COMPLETED
        await _bg(candidate.save)(update_fields=[
            "status", "call_duration", "call_outcome", "interest_level",
            "transcript", "notes", "ai_summary", "ended_at",
        ])
        logger.info("[Campaign] Synced %s → outcome=%s", candidate.name, candidate.call_outcome)
    except Exception as exc:
        logger.error("[Campaign] Sync error candidate_id=%d: %s", candidate_id, exc)


# ---------------------------------------------------------------------------
# Startup auto-resume — re-launch tasks for campaigns that were RUNNING when
# the container last restarted (their asyncio tasks were wiped from memory).
# ---------------------------------------------------------------------------

async def resume_running_campaigns() -> None:
    """Called once on ASGI startup. Re-creates tasks for any RUNNING campaigns."""
    from .models import Campaign, CampaignCandidate
    try:
        running_ids = await _bg(
            lambda: list(Campaign.objects.filter(status=Campaign.RUNNING).values_list("id", flat=True))
        )()
        for cid in running_ids:
            existing = _CAMPAIGN_TASKS.get(cid)
            if existing and not existing.done():
                continue  # Already running (shouldn't happen on fresh start, but be safe)
            # Reset any CALLING candidates stuck from the last interrupted run
            reset = await _bg(
                CampaignCandidate.objects.filter(
                    campaign_id=cid, status=CampaignCandidate.CALLING,
                ).update
            )(status=CampaignCandidate.PENDING)
            logger.info("[Startup] Resuming campaign %d (reset %d CALLING→PENDING)", cid, reset)
            task = asyncio.create_task(_run_campaign_caller(cid))
            _CAMPAIGN_TASKS[cid] = task
    except Exception as exc:
        logger.error("[Startup] Failed to resume running campaigns: %s", exc)


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def _safe_int(val, default: int) -> int:
    try:
        return int(val)
    except (TypeError, ValueError):
        return default
