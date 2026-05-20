"""Vox AI — FastAPI backend (SQLAlchemy async + redis.asyncio + Uvicorn)."""
from __future__ import annotations

import asyncio
import io
import json
import logging
import os
import re
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timezone as _tz
from typing import Any

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response, StreamingResponse
from sqlalchemy import select, update
from twilio.rest import Client

from cache import cache
from config import settings
from database import AsyncSessionLocal, Base, engine
from models import CallSession, Campaign, CampaignCandidate

logger = logging.getLogger(__name__)

_E164_RE = re.compile(r"^\+[1-9]\d{6,14}$")
_SANITIZE_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_MAX_JD_LENGTH = 4000
_MAX_NAME_LENGTH = 100
_MAX_RESUME_LENGTH = 8000
_MAX_RESUME_BYTES = 5 * 1024 * 1024
_ALLOWED_RESUME_MIME = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}

_BACKGROUND_TASKS: set = set()
_CAMPAIGN_TASKS: dict[int, asyncio.Task] = {}

_UTC = _tz.utc


def _now() -> datetime:
    return datetime.now(_UTC)


# ---------------------------------------------------------------------------
# App lifespan
# ---------------------------------------------------------------------------

@asynccontextmanager
async def lifespan(app: FastAPI):
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s %(message)s")

    if settings.SENTRY_DSN:
        import sentry_sdk
        sentry_sdk.init(dsn=settings.SENTRY_DSN, traces_sample_rate=0.1)

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    await resume_running_campaigns()
    yield


app = FastAPI(title="Vox AI", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.allowed_origins_list(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _validate_e164(number: str) -> bool:
    return bool(_E164_RE.match(number.strip()))


def _sanitize(text: str, max_length: int = 4000) -> str:
    return _SANITIZE_RE.sub("", text).strip()[:max_length]


def _clean_host(host_url: str) -> str:
    return (
        host_url.strip()
        .replace("https://", "")
        .replace("http://", "")
        .replace("wss://", "")
        .replace("ws://", "")
        .rstrip("/")
    )


def _build_twiml(stream_url: str) -> str:
    escaped = stream_url.replace("&", "&amp;")
    return (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        "<Response>\n"
        "    <Connect>\n"
        f'        <Stream url="{escaped}" />\n'
        "    </Connect>\n"
        "</Response>"
    )


async def _rate_limit(key: str, max_calls: int, window_seconds: int) -> bool:
    slot = int(time.time() // window_seconds)
    cache_key = f"vox:rl:{key}:{slot}"
    await cache.add(cache_key, 0, timeout=window_seconds * 2)
    count = await cache.incr(cache_key)
    return count > max_calls


def _get_client_ip(request: Request) -> str:
    xff = request.headers.get("x-forwarded-for", "")
    if xff:
        ips = [ip.strip() for ip in xff.split(",") if ip.strip()]
        if ips:
            return ips[-1]
    return request.client.host if request.client else "unknown"


def _check_api_key(request: Request) -> bool:
    required = settings.VOX_API_KEY.strip()
    if not required:
        return True
    provided = (
        request.headers.get("x-vox-api-key", "")
        or request.query_params.get("api_key", "")
    ).strip()
    return provided == required


async def _is_call_already_active(phone: str) -> bool:
    return bool(await cache.get(f"vox:active_call:{phone}"))


async def _mark_call_active(phone: str) -> None:
    await cache.set(f"vox:active_call:{phone}", 1, timeout=60)


def _verify_twilio_signature(url: str, params: dict, signature: str) -> bool:
    auth_token = settings.TWILIO_AUTH_TOKEN.strip()
    if not auth_token:
        if settings.DEBUG:
            logger.warning("[Security] TWILIO_AUTH_TOKEN not set — skipping check (dev)")
            return True
        logger.error("[Security] TWILIO_AUTH_TOKEN missing in production")
        return False
    try:
        from twilio.request_validator import RequestValidator
        return RequestValidator(auth_token).validate(url, params, signature)
    except Exception as exc:
        logger.error("[Security] Twilio signature check error: %s", exc)
        return False


# ---------------------------------------------------------------------------
# _place_call — core Twilio call placement (async)
# ---------------------------------------------------------------------------

async def _place_call(
    *,
    to_number: str,
    from_number: str,
    host_url: str,
    jd: str = "Software Engineer role",
    name: str = "Candidate",
    resume_text: str = "",
    retry_num: int = 0,
    prior_transcript: list | None = None,
    prior_notes: dict | None = None,
    recruiter_inputs: dict | None = None,
    voice_id: str = "",
) -> dict:
    from chat.agent import DEFAULT_VOICE_ID, VOICE_PROFILES
    voice_id = voice_id if voice_id in VOICE_PROFILES else DEFAULT_VOICE_ID

    clean_host = _clean_host(host_url)
    token = str(uuid.uuid4())

    session_data: dict = {
        "jd": jd, "name": name, "phone": to_number,
        "retry_num": retry_num, "voice_id": voice_id,
    }
    if resume_text:
        session_data["resume_text"] = resume_text
    if prior_transcript:
        session_data["prior_transcript"] = prior_transcript
    if prior_notes:
        session_data["prior_notes"] = prior_notes
    if recruiter_inputs:
        session_data["recruiter_inputs"] = recruiter_inputs

    stream_url = f"wss://{clean_host}/media-stream/{token}"
    twiml = _build_twiml(stream_url)

    account_sid = settings.TWILIO_ACCOUNT_SID.strip()
    auth_token = settings.TWILIO_AUTH_TOKEN.strip()
    if not (account_sid and auth_token):
        raise RuntimeError("Twilio credentials not configured")

    loop = asyncio.get_running_loop()
    client = Client(account_sid, auth_token)
    status_callback_url = f"https://{clean_host}/api/call-status/"
    call = await loop.run_in_executor(
        None,
        lambda: client.calls.create(
            twiml=twiml,
            to=to_number,
            from_=from_number,
            status_callback=status_callback_url,
            status_callback_method="POST",
        ),
    )

    await _mark_call_active(to_number)
    session_data["call_sid"] = call.sid
    await cache.set(f"vox:{token}", session_data, timeout=3600)
    await cache.set(f"vox:call:{call.sid}", {
        "phone": to_number, "name": name, "jd": jd,
        "resume_text": resume_text, "retry_num": retry_num,
        "host_url": host_url, "from_number": from_number,
        "prior_transcript": prior_transcript or [],
        "prior_notes": prior_notes or {},
        "recruiter_inputs": recruiter_inputs or {},
        "voice_id": voice_id,
    }, timeout=3600)

    db_session_data = {k: v for k, v in session_data.items() if k != "resume_text"}
    try:
        async with AsyncSessionLocal() as db:
            db.add(CallSession(
                call_sid=call.sid,
                candidate_name=name,
                candidate_phone=to_number,
                job_description=jd,
                resume_text=resume_text,
                call_channel="twilio",
                session_token=token,
                session_data=db_session_data,
            ))
            await db.commit()
    except Exception as _db_err:
        logger.warning("[DB] Failed to pre-create CallSession: %s", _db_err)

    logger.info("[Call] TwiML stream_url=%s", stream_url)
    return {"status": "Call initiated", "call_sid": call.sid, "stream_url": stream_url, "token": token}


# ---------------------------------------------------------------------------
# Campaign result sync helpers (used by _sync_campaign_candidate in agent.py)
# ---------------------------------------------------------------------------

async def _sync_call_result(candidate_id: int, call_sid: str) -> None:
    try:
        async with AsyncSessionLocal() as db:
            res = await db.execute(select(CallSession).where(CallSession.call_sid == call_sid))
            session = res.scalars().first()
            cand_res = await db.execute(
                select(CampaignCandidate).where(CampaignCandidate.id == candidate_id)
            )
            candidate = cand_res.scalars().first()

            if not candidate:
                return
            if not session:
                candidate.status = CampaignCandidate.FAILED
                await db.commit()
                return

            if session.created_at and session.ended_at:
                candidate.call_duration = int((session.ended_at - session.created_at).total_seconds())

            candidate.call_outcome = session.call_outcome or ""
            candidate.interest_level = session.call_outcome or ""
            candidate.transcript = session.transcript or []
            candidate.notes = session.notes or {}
            candidate.ended_at = session.ended_at
            if session.notes:
                bullets = (session.notes or {}).get("summary_bullets", [])
                candidate.ai_summary = " | ".join(bullets[:3]) if bullets else ""
            candidate.status = CampaignCandidate.COMPLETED
            await db.commit()
            logger.info("[Campaign] Synced %s → outcome=%s", candidate.name, candidate.call_outcome)
    except Exception as exc:
        logger.error("[Campaign] Sync error candidate_id=%d: %s", candidate_id, exc)


# ---------------------------------------------------------------------------
# Pre-cache interview context
# ---------------------------------------------------------------------------

async def _precache_interview_context(token: str, jd: str, name: str, recruiter_inputs: dict | None) -> None:
    from chat.agents.manager import AgentManager
    try:
        manager = AgentManager(session_id=f"precache:{token}")
        context = await manager.prepare_session(jd=jd, candidate_name=name, recruiter_inputs=recruiter_inputs)
        await cache.set(f"vox:context:{token}", context.to_dict(), timeout=3600)
        logger.info("[Precache] JD context ready token=%s status=%s", token, context.recruiter_status)
    except Exception as exc:
        logger.warning("[Precache] Failed for token=%s: %s", token, exc)


# ---------------------------------------------------------------------------
# HTTP routes
# ---------------------------------------------------------------------------

@app.get("/")
@app.get("/health/")
async def healthcheck():
    return {"status": "ok"}


@app.post("/api/call/")
async def initiate_call(request: Request):
    client_ip = _get_client_ip(request)
    if await _rate_limit(f"call:{client_ip}", max_calls=10, window_seconds=60):
        raise HTTPException(status_code=429, detail="Too many requests — please wait a moment")

    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    to_number = (data.get("phone") or data.get("to_number") or "").strip()
    raw_jd = data.get("jd") or "Software Engineer role"
    raw_name = data.get("name") or "Candidate"

    if not to_number:
        raise HTTPException(status_code=400, detail="phone is required")
    if not _validate_e164(to_number):
        raise HTTPException(status_code=400, detail="phone must be in E.164 format (e.g. +919876543210)")

    jd = _sanitize(raw_jd, _MAX_JD_LENGTH)
    name = _sanitize(raw_name, _MAX_NAME_LENGTH)
    resume_text = _sanitize(data.get("resume_text") or "", _MAX_RESUME_LENGTH)
    if not jd:
        raise HTTPException(status_code=400, detail="jd cannot be empty")

    _RECRUITER_FIELDS = {
        "company_overview", "team_details", "company_location",
        "years_of_experience", "ctc_range", "required_joining_timeline", "work_location_type",
    }
    raw_ri = data.get("recruiter_inputs") or {}
    recruiter_inputs = {
        k: _sanitize(str(v), 500)
        for k, v in raw_ri.items()
        if k in _RECRUITER_FIELDS and isinstance(v, str) and v.strip()
    } or None

    public_url = (data.get("host_url") or settings.PUBLIC_URL).strip()
    if not public_url or "ngrok_url_here" in public_url:
        raise HTTPException(status_code=400, detail="PUBLIC_URL not configured in .env")

    from_number = (data.get("from_number") or settings.TWILIO_PHONE_NUMBER).strip()
    if not from_number:
        raise HTTPException(status_code=500, detail="TWILIO_PHONE_NUMBER not configured")

    if await _is_call_already_active(to_number):
        raise HTTPException(status_code=409, detail="A call to this number is already in progress")

    raw_voice_id = _sanitize(data.get("voice_id") or "", 20)

    try:
        result = await _place_call(
            to_number=to_number, from_number=from_number, host_url=public_url,
            jd=jd, name=name, resume_text=resume_text,
            recruiter_inputs=recruiter_inputs, voice_id=raw_voice_id,
        )
    except Exception as e:
        logger.error("[Call-Error] %s", e)
        raise HTTPException(status_code=500, detail="Failed to initiate call")

    t = asyncio.create_task(_precache_interview_context(result["token"], jd, name, recruiter_inputs))
    _BACKGROUND_TASKS.add(t)
    t.add_done_callback(_BACKGROUND_TASKS.discard)

    return {"status": "success", "call_sid": result["call_sid"]}


@app.post("/outgoing-call/")
async def outgoing_call(request: Request):
    if not _check_api_key(request):
        raise HTTPException(status_code=401, detail="Unauthorized")
    if await _rate_limit(f"call:{_get_client_ip(request)}", max_calls=10, window_seconds=60):
        raise HTTPException(status_code=429, detail="Too many requests")

    try:
        body = await request.json()
    except Exception:
        body = {}

    to_number = (request.query_params.get("to_number") or body.get("to_number") or body.get("phone") or "").strip()
    from_number = (request.query_params.get("from_number") or body.get("from_number") or settings.TWILIO_PHONE_NUMBER).strip()
    host_url = (request.query_params.get("host_url") or body.get("host_url") or settings.PUBLIC_URL).strip()

    if not to_number or not _validate_e164(to_number):
        raise HTTPException(status_code=400, detail="to_number must be E.164")
    if not from_number:
        raise HTTPException(status_code=400, detail="from_number is required")
    if not host_url or "ngrok_url_here" in host_url:
        raise HTTPException(status_code=400, detail="host_url / PUBLIC_URL not configured")

    try:
        result = await _place_call(to_number=to_number, from_number=from_number, host_url=host_url)
        return result
    except Exception as e:
        logger.error("[Call-Error] %s", e)
        raise HTTPException(status_code=500, detail="Failed to initiate call")


@app.post("/api/call-status/")
async def call_status_webhook(request: Request):
    from chat.retry_manager import CallRetryManager, RETRY_1_DELAY, RETRY_2_DELAY

    form = await request.form()
    params = dict(form)

    public_url = settings.PUBLIC_URL.strip().rstrip("/")
    path = request.url.path.lstrip("/")
    url = f"https://{_clean_host(public_url)}/{path}" if public_url else str(request.url)
    sig = request.headers.get("x-twilio-signature", "")
    if not _verify_twilio_signature(url, params, sig):
        return Response(status_code=403)

    call_sid = params.get("CallSid", "")
    call_status = params.get("CallStatus", "")
    logger.info("[CallStatus] sid=%s status=%s", call_sid, call_status)

    TERMINAL = {"completed", "no-answer", "busy", "failed", "canceled"}
    if call_status in TERMINAL and call_sid:
        campaign_call = await cache.get(f"vox:campaign_call:{call_sid}")
        if campaign_call:
            cand_id = campaign_call.get("candidate_id")
            if cand_id:
                try:
                    failed = {"no-answer", "busy", "failed", "canceled"}
                    new_st = CampaignCandidate.FAILED if call_status in failed else CampaignCandidate.COMPLETED
                    async with AsyncSessionLocal() as db:
                        await db.execute(
                            update(CampaignCandidate)
                            .where(CampaignCandidate.id == cand_id)
                            .values(status=new_st)
                        )
                        await db.commit()
                except Exception as _ce:
                    logger.warning("[CallStatus] Failed to update campaign candidate: %s", _ce)

    if call_status == "completed" and call_sid:
        try:
            async with AsyncSessionLocal() as db:
                await db.execute(
                    update(CallSession)
                    .where(CallSession.call_sid == call_sid, CallSession.ended_at == None)
                    .values(ended_at=_now(), call_outcome="COMPLETED")
                )
                await db.commit()
        except Exception as _ce:
            logger.warning("[CallStatus] Failed to stamp ended_at: %s", _ce)

    if call_status not in ("no-answer", "busy", "failed"):
        return Response(status_code=204)

    if call_sid:
        try:
            outcome_map = {"no-answer": "BUSY", "busy": "BUSY", "failed": "FAILED"}
            outcome = outcome_map.get(call_status, "BUSY")
            async with AsyncSessionLocal() as db:
                result = await db.execute(
                    update(CallSession)
                    .where(CallSession.call_sid == call_sid, CallSession.ended_at == None)
                    .values(ended_at=_now(), call_outcome=outcome)
                    .returning(CallSession.id)
                )
                await db.commit()
                if not result.fetchone():
                    logger.info("[CallStatus] Duplicate webhook sid=%s — already processed", call_sid)
                    return Response(status_code=204)
        except Exception as _db_err:
            logger.warning("[CallStatus] Failed to mark session ended: %s", _db_err)

    session = await cache.get(f"vox:call:{call_sid}")
    if not session:
        logger.warning("[CallStatus] No session data for CallSid=%s", call_sid)
        return Response(status_code=500)

    phone = session.get("phone", "")
    name = session.get("name", "Candidate")
    jd = session.get("jd", "")
    resume_text = session.get("resume_text", "")
    retry_num = session.get("retry_num", 0)
    host_url = session.get("host_url", "") or settings.PUBLIC_URL
    from_number = session.get("from_number", "") or settings.TWILIO_PHONE_NUMBER
    prior_transcript = session.get("prior_transcript", [])
    prior_notes = session.get("prior_notes", {})
    recruiter_inputs = session.get("recruiter_inputs") or None
    voice_id = session.get("voice_id", "")

    if not phone or not host_url or not from_number:
        return Response(status_code=204)

    if retry_num >= CallRetryManager.MAX_RETRIES:
        await CallRetryManager.clear(phone)
        return Response(status_code=204)

    new_retry_num = await CallRetryManager.record_drop(
        phone=phone, name=name, jd=jd, transcript=prior_transcript,
        notes=prior_notes, resume_text=resume_text,
        recruiter_inputs=recruiter_inputs, voice_id=voice_id,
    )

    delay = RETRY_1_DELAY if new_retry_num == 1 else RETRY_2_DELAY
    state = await CallRetryManager.load(phone)
    CallRetryManager.schedule_callback(
        phone=phone, name=name, jd=jd,
        transcript=state.get("transcript", prior_transcript),
        notes=state.get("notes", prior_notes),
        resume_text=state.get("resume_text", resume_text),
        recruiter_inputs=state.get("recruiter_inputs") or recruiter_inputs,
        voice_id=state.get("voice_id", voice_id),
        retry_num=new_retry_num, delay_seconds=delay,
        host_url=host_url, from_number=from_number,
    )
    return Response(status_code=204)


@app.get("/api/voices/")
async def list_voices():
    from chat.agent import VOICE_PROFILES
    return {"voices": [
        {"id": v["id"], "display_name": v["display_name"], "accent": v["accent"], "description": v["description"]}
        for v in VOICE_PROFILES.values()
    ]}


@app.post("/api/upload-resume/")
async def upload_resume(request: Request, resume: UploadFile = File(...)):
    if await _rate_limit(f"resume:{_get_client_ip(request)}", max_calls=5, window_seconds=60):
        raise HTTPException(status_code=429, detail="Too many uploads — please wait a moment")

    if resume.size and resume.size > _MAX_RESUME_BYTES:
        raise HTTPException(status_code=400, detail="File too large (max 5 MB)")

    fname = (resume.filename or "").lower()
    content_type = resume.content_type or ""
    is_pdf = fname.endswith(".pdf")
    is_docx = fname.endswith(".docx")
    if not (is_pdf or is_docx):
        raise HTTPException(status_code=400, detail="Only PDF and DOCX files are supported")
    if content_type and content_type not in _ALLOWED_RESUME_MIME and "octet-stream" not in content_type:
        raise HTTPException(status_code=400, detail="File content-type does not match extension")

    raw = await resume.read()
    if len(raw) > _MAX_RESUME_BYTES:
        raise HTTPException(status_code=400, detail="File too large (max 5 MB)")

    try:
        if is_pdf:
            from pypdf import PdfReader
            text = "\n".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(raw)).pages).strip()
        else:
            from docx import Document
            text = "\n".join(p.text for p in Document(io.BytesIO(raw)).paragraphs if p.text.strip()).strip()
    except Exception as e:
        raise HTTPException(status_code=400, detail="Could not parse file — try re-saving as PDF")

    if not text.strip():
        raise HTTPException(status_code=400, detail="No text found — file may be a scanned image")

    capped = text[:_MAX_RESUME_LENGTH]
    return {"status": "success", "text": capped, "chars": len(capped)}


@app.get("/api/sessions/")
async def list_sessions(request: Request, limit: int = 50, offset: int = 0):
    if not _check_api_key(request):
        raise HTTPException(status_code=401, detail="Unauthorized")

    limit = min(limit, 200)
    offset = max(offset, 0)

    try:
        from sqlalchemy import func
        async with AsyncSessionLocal() as db:
            total = (await db.execute(select(func.count()).select_from(CallSession))).scalar()
            res = await db.execute(
                select(CallSession).order_by(CallSession.created_at.desc()).offset(offset).limit(limit)
            )
            sessions = res.scalars().all()
    except Exception:
        raise HTTPException(status_code=500, detail="DB unavailable")

    outcome_counts: dict[str, int] = {}
    compat_counts: dict[str, int] = {}
    total_score = 0; score_count = 0
    total_conf = 0.0; conf_count = 0
    payload = []

    for s in sessions:
        oc = s.call_outcome or "unknown"
        outcome_counts[oc] = outcome_counts.get(oc, 0) + 1
        compat = (
            (s.candidate_summary or {}).get("compatibility_level")
            if isinstance(s.candidate_summary, dict) else None
        ) or "unknown"
        compat_counts[compat] = compat_counts.get(compat, 0) + 1
        if s.intent_score is not None:
            total_score += s.intent_score; score_count += 1
        if s.eval_confidence is not None:
            total_conf += s.eval_confidence; conf_count += 1
        payload.append({
            "id": s.id, "call_sid": s.call_sid,
            "candidate_name": s.candidate_name, "candidate_phone": s.candidate_phone,
            "job_description": s.job_description, "resume_text": s.resume_text,
            "summary": s.summary, "intent_score": s.intent_score,
            "call_outcome": s.call_outcome, "call_channel": s.call_channel,
            "created_at": s.created_at.isoformat() if s.created_at else None,
            "ended_at": s.ended_at.isoformat() if s.ended_at else None,
            "notes": s.notes, "candidate_summary": s.candidate_summary,
            "eval_confidence": s.eval_confidence, "dimension_scores": s.dimension_scores,
            "eval_reasoning": s.eval_reasoning,
            "transcript_length": len(s.transcript or []),
        })

    return {
        "status": "success", "total_sessions": total,
        "limit": limit, "offset": offset,
        "outcome_counts": outcome_counts, "compatibility_counts": compat_counts,
        "average_score": round(total_score / score_count, 1) if score_count else None,
        "average_confidence": round(total_conf / conf_count, 2) if conf_count else None,
        "sessions": payload,
    }


@app.get("/api/session/{call_sid}/")
async def session_status(call_sid: str, request: Request):
    if not _check_api_key(request):
        raise HTTPException(status_code=401, detail="Unauthorized")

    try:
        async with AsyncSessionLocal() as db:
            res = await db.execute(
                select(CallSession).where(CallSession.call_sid == call_sid)
                .order_by(CallSession.created_at.desc())
            )
            session = res.scalars().first()
    except Exception:
        raise HTTPException(status_code=500, detail="DB unavailable")

    if not session:
        raise HTTPException(status_code=404, detail="Not found")

    if not session.ended_at:
        if await cache.get(f"vox:ended:{call_sid}"):
            return {"status": "evaluating"}
        return JSONResponse({"status": "pending"}, status_code=202)

    return {
        "status": "complete",
        "score": session.intent_score,
        "call_outcome": session.call_outcome,
        "notes": session.notes,
        "candidate_summary": session.candidate_summary,
        "eval_confidence": session.eval_confidence,
    }


@app.post("/api/call/{call_sid}/end/")
async def end_call(call_sid: str):
    account_sid = settings.TWILIO_ACCOUNT_SID.strip()
    auth_token = settings.TWILIO_AUTH_TOKEN.strip()
    if not (account_sid and auth_token):
        raise HTTPException(status_code=500, detail="Twilio not configured")
    try:
        loop = asyncio.get_running_loop()
        client = Client(account_sid, auth_token)
        await loop.run_in_executor(None, lambda: client.calls(call_sid).update(status="completed"))
        return {"status": "ok"}
    except Exception as e:
        err = str(e)
        if "20404" in err or "not found" in err.lower():
            raise HTTPException(status_code=404, detail="Call not found")
        raise HTTPException(status_code=500, detail="Failed to end call")


# ---------------------------------------------------------------------------
# Campaign endpoints
# ---------------------------------------------------------------------------

@app.get("/api/campaigns/")
async def list_campaigns():
    async with AsyncSessionLocal() as db:
        res = await db.execute(select(Campaign).order_by(Campaign.created_at.desc()))
        campaigns = res.scalars().all()

    campaign_ids = [c.id for c in campaigns]
    cands_by_campaign: dict[int, list] = {c.id: [] for c in campaigns}
    if campaign_ids:
        async with AsyncSessionLocal() as db:
            r = await db.execute(
                select(CampaignCandidate).where(
                    CampaignCandidate.campaign_id.in_(campaign_ids),
                    CampaignCandidate.is_valid == True,
                    CampaignCandidate.is_duplicate == False,
                )
            )
            for cand in r.scalars().all():
                cands_by_campaign[cand.campaign_id].append(cand)

    return {"campaigns": [{
        "id": c.id, "name": c.name, "status": c.status,
        "voice_id": c.voice_id, "total_uploaded": c.total_uploaded,
        "valid_count": c.valid_count, "stats": _calc_stats(cands_by_campaign[c.id]),
        "created_at": c.created_at.isoformat(),
        "started_at": c.started_at.isoformat() if c.started_at else None,
    } for c in campaigns]}


@app.post("/api/campaigns/", status_code=201)
async def create_campaign(
    request: Request,
    file: UploadFile = File(...),
    campaign_name: str = Form(default=""),
    job_description: str = Form(default=""),
    voice_id: str = Form(default="priya"),
    delay_seconds: int = Form(default=30),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")

    if not campaign_name.strip():
        campaign_name = f"Campaign {_now():%Y-%m-%d %H:%M}"

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}")

    if not all_rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    header = [str(h or "").strip().lower() for h in all_rows[0]]
    name_col = next((i for i, h in enumerate(header) if "name" in h), None)
    phone_col = next(
        (i for i, h in enumerate(header) if any(k in h for k in ("phone", "mobile", "number", "contact", "tel"))),
        None,
    )
    if name_col is None or phone_col is None:
        raise HTTPException(status_code=400, detail="Could not find 'Name' and 'Phone' columns")

    data_rows = []
    for row in all_rows[1:]:
        if not row:
            continue
        n_val = row[name_col] if len(row) > name_col else None
        p_val = row[phone_col] if len(row) > phone_col else None
        data_rows.append({"name": n_val, "phone": p_val})

    result = _validate_candidates(data_rows)
    total = len(data_rows)

    async with AsyncSessionLocal() as db:
        campaign = Campaign(
            name=campaign_name, job_description=job_description,
            voice_id=voice_id, delay_seconds=delay_seconds,
            total_uploaded=total,
            valid_count=len(result["valid"]),
            invalid_count=len(result["invalid"]),
            duplicate_count=len(result["duplicates"]),
        )
        db.add(campaign)
        await db.flush()

        to_create = [
            CampaignCandidate(campaign_id=campaign.id, name=c["name"], phone=c["phone"],
                              is_valid=True, is_duplicate=False)
            for c in result["valid"]
        ] + [
            CampaignCandidate(campaign_id=campaign.id, name=c.get("name", ""),
                              phone=c.get("raw_phone", c.get("phone", "")),
                              is_valid=False, is_duplicate=False,
                              validation_error=c.get("error", ""))
            for c in result["invalid"]
        ] + [
            CampaignCandidate(campaign_id=campaign.id, name=c.get("name", ""),
                              phone=c.get("raw_phone", c.get("phone", "")),
                              is_valid=False, is_duplicate=True, validation_error="Duplicate")
            for c in result["duplicates"]
        ]
        db.add_all(to_create)
        await db.commit()
        campaign_id = campaign.id
        campaign_name_out = campaign.name

    return {
        "campaign_id": campaign_id, "campaign_name": campaign_name_out,
        "validation": {
            "total_uploaded": total, "valid": len(result["valid"]),
            "invalid": len(result["invalid"]), "duplicates": len(result["duplicates"]),
            "invalid_details": [
                {"name": c.get("name"), "phone": c.get("raw_phone", ""), "error": c.get("error")}
                for c in result["invalid"]
            ],
            "duplicate_details": [{"name": c.get("name"), "phone": c.get("raw_phone", "")} for c in result["duplicates"]],
        },
        "candidates": [{"name": c["name"], "phone": c["phone"]} for c in result["valid"]],
    }


@app.get("/api/campaigns/{campaign_id}/")
async def get_campaign(campaign_id: int):
    async with AsyncSessionLocal() as db:
        res = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
        campaign = res.scalars().first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if campaign.status == Campaign.RUNNING:
        existing = _CAMPAIGN_TASKS.get(campaign_id)
        if not existing or existing.done():
            async with AsyncSessionLocal() as db:
                await db.execute(
                    update(CampaignCandidate)
                    .where(CampaignCandidate.campaign_id == campaign_id,
                           CampaignCandidate.status == CampaignCandidate.CALLING)
                    .values(status=CampaignCandidate.PENDING)
                )
                await db.commit()
            task = asyncio.create_task(_run_campaign_caller(campaign_id))
            _CAMPAIGN_TASKS[campaign_id] = task
            logger.info("[Campaign-%d] Auto-resumed via GET poll", campaign_id)

    async with AsyncSessionLocal() as db:
        res = await db.execute(
            select(CampaignCandidate)
            .where(CampaignCandidate.campaign_id == campaign_id,
                   CampaignCandidate.is_valid == True,
                   CampaignCandidate.is_duplicate == False)
            .order_by(CampaignCandidate.created_at)
        )
        cands = res.scalars().all()

    stats = _calc_stats(cands)
    candidates = [{
        "id": c.id, "name": c.name, "phone": c.phone, "status": c.status,
        "call_outcome": c.call_outcome, "interest_level": c.interest_level,
        "call_duration": c.call_duration, "ai_summary": c.ai_summary,
        "called_at": c.called_at.isoformat() if c.called_at else None,
        "ended_at": c.ended_at.isoformat() if c.ended_at else None,
    } for c in cands]

    return {
        "id": campaign.id, "name": campaign.name, "status": campaign.status,
        "voice_id": campaign.voice_id, "delay_seconds": campaign.delay_seconds,
        "job_description": campaign.job_description,
        "validation_summary": {
            "total_uploaded": campaign.total_uploaded, "valid": campaign.valid_count,
            "invalid": campaign.invalid_count, "duplicates": campaign.duplicate_count,
        },
        "stats": stats, "candidates": candidates,
        "created_at": campaign.created_at.isoformat(),
        "started_at": campaign.started_at.isoformat() if campaign.started_at else None,
        "completed_at": campaign.completed_at.isoformat() if campaign.completed_at else None,
    }


@app.post("/api/campaigns/{campaign_id}/start/")
async def start_campaign(campaign_id: int):
    async with AsyncSessionLocal() as db:
        res = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
        campaign = res.scalars().first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if campaign.status == Campaign.RUNNING:
        existing = _CAMPAIGN_TASKS.get(campaign_id)
        if existing and not existing.done():
            raise HTTPException(status_code=400, detail="Campaign is already running")

    if campaign.valid_count == 0:
        raise HTTPException(status_code=400, detail="No valid candidates to call")

    async with AsyncSessionLocal() as db:
        await db.execute(
            update(Campaign).where(Campaign.id == campaign_id)
            .values(status=Campaign.RUNNING,
                    started_at=campaign.started_at or _now())
        )
        await db.execute(
            update(CampaignCandidate)
            .where(CampaignCandidate.campaign_id == campaign_id,
                   CampaignCandidate.status == CampaignCandidate.CALLING)
            .values(status=CampaignCandidate.PENDING)
        )
        await db.commit()

    old = _CAMPAIGN_TASKS.get(campaign_id)
    if old and not old.done():
        old.cancel()

    task = asyncio.create_task(_run_campaign_caller(campaign_id))
    _CAMPAIGN_TASKS[campaign_id] = task
    return {"status": "started", "campaign_id": campaign_id}


@app.post("/api/campaigns/{campaign_id}/pause/")
async def pause_campaign(campaign_id: int):
    async with AsyncSessionLocal() as db:
        res = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
        campaign = res.scalars().first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    async with AsyncSessionLocal() as db:
        await db.execute(update(Campaign).where(Campaign.id == campaign_id).values(status=Campaign.PAUSED))
        await db.commit()

    task = _CAMPAIGN_TASKS.get(campaign_id)
    if task and not task.done():
        task.cancel()
    return {"status": "paused"}


@app.get("/api/campaigns/{campaign_id}/export/")
async def export_campaign(campaign_id: int):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    async with AsyncSessionLocal() as db:
        res = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
        campaign = res.scalars().first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    async with AsyncSessionLocal() as db:
        res = await db.execute(
            select(CampaignCandidate)
            .where(CampaignCandidate.campaign_id == campaign_id,
                   CampaignCandidate.is_valid == True,
                   CampaignCandidate.is_duplicate == False)
            .order_by(CampaignCandidate.created_at)
        )
        cands = res.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screening Report"
    headers = ["#", "Name", "Phone", "Status", "Call Outcome", "Interest",
               "Duration (min)", "AI Summary", "Called At", "Ended At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
        cell.alignment = Alignment(horizontal="center")

    STATUS_COLOURS = {"completed": "D4EDDA", "failed": "F8D7DA", "calling": "FFF3CD", "pending": "E2E3E5"}
    for i, c in enumerate(cands, 1):
        dur = round(c.call_duration / 60, 1) if c.call_duration else ""
        ws.append([
            i, c.name, c.phone, c.status.upper(), c.call_outcome or "",
            c.interest_level or "", dur, c.ai_summary or "",
            c.called_at.strftime("%Y-%m-%d %H:%M") if c.called_at else "",
            c.ended_at.strftime("%Y-%m-%d %H:%M") if c.ended_at else "",
        ])
        colour = STATUS_COLOURS.get(c.status, "FFFFFF")
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor=colour)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in col) + 4, 60
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"campaign_{campaign.name.replace(' ', '_')}_{_now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# WebSocket — Twilio Media Stream
# ---------------------------------------------------------------------------

class _FinalizeConsumer:
    """Minimal consumer bridge passed to GeminiLiveBridge.finalize_consumer."""
    async def send_recap(self, score, reason) -> None:
        logger.info("[WS] Session finalized — score=%s", score)


async def _twilio_websocket_handler(websocket: WebSocket, token: str):
    _CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
    _DEFAULT_JD = "Software Engineer at a high-growth startup."

    session: dict = {}
    if token:
        session = await cache.get(f"vox:{token}") or {}
        if not session:
            try:
                async with AsyncSessionLocal() as db:
                    res = await db.execute(
                        select(CallSession).where(CallSession.session_token == token)
                    )
                    db_row = res.scalars().first()
                    if db_row and db_row.session_data:
                        session = dict(db_row.session_data)
                        if db_row.resume_text and "resume_text" not in session:
                            session["resume_text"] = db_row.resume_text
                        logger.info("[WS] Redis miss — recovered from DB token=%s", token[:12])
            except Exception as exc:
                logger.error("[WS] DB fallback failed: %s", exc)

    if not session:
        logger.warning("[WS] Invalid/expired token=%s — rejecting", token[:12])
        await websocket.accept()
        await websocket.close(code=4401)
        return

    await websocket.accept()

    def _clean(s: str, max_len: int = 200) -> str:
        return _CONTROL_RE.sub("", (s or "").strip())[:max_len]

    from chat.agent import DEFAULT_VOICE_ID, VOICE_PROFILES, build_enriched_system_prompt, build_vox_greeting_kickoff
    from chat.agents.manager import AgentManager
    from chat.agents.schemas import InterviewContext
    from chat.gemini_recruiter import GeminiLiveBridge, RECRUITER_PROMPT, SARAH_GREETING_KICKOFF
    from chat.retry_manager import CallRetryManager, RETRY_1_DELAY, RETRY_2_DELAY

    name = _clean(session.get("name") or "Candidate", 100)
    jd = _clean(session.get("jd") or "", 4000)
    phone = _clean(session.get("phone") or "", 20)
    resume_text = _clean(session.get("resume_text") or "", 8000)
    raw_voice = _clean(session.get("voice_id") or DEFAULT_VOICE_ID, 20)
    voice_id = raw_voice if raw_voice in VOICE_PROFILES else DEFAULT_VOICE_ID
    voice_profile = VOICE_PROFILES[voice_id]
    call_sid = _clean(session.get("call_sid") or "", 50)
    retry_num = int(session.get("retry_num", 0))
    prior_transcript = session.get("prior_transcript", [])
    prior_notes = session.get("prior_notes", {})
    recruiter_inputs = session.get("recruiter_inputs") or None

    if not jd or len(jd.strip()) < 20:
        logger.warning("[WS] No usable JD — using default")
        jd = _DEFAULT_JD

    context: InterviewContext | None = None
    cached_ctx = await cache.get(f"vox:context:{token}")
    if cached_ctx:
        context = InterviewContext.from_dict(cached_ctx)
        logger.info("[WS] Pre-cached context loaded status=%s", context.recruiter_status)

    if context is None:
        session_id = str(uuid.uuid4())
        manager = AgentManager(session_id=session_id)
        context = await manager.prepare_session(jd=jd, candidate_name=name, recruiter_inputs=recruiter_inputs)

    if jd:
        system_prompt = build_enriched_system_prompt(name, jd, context, resume_text=resume_text, voice_profile=voice_profile)
    else:
        system_prompt = RECRUITER_PROMPT.strip()

    if retry_num > 0 and prior_transcript:
        system_prompt += CallRetryManager.build_continuity_section(prior_transcript, prior_notes)

    if retry_num > 0:
        reconnect = CallRetryManager.reconnect_greeting(name, retry_num)
        greeting_kickoff = (
            f"Open with EXACTLY this line and nothing else first: "
            f'"{reconnect}" — then naturally continue the conversation '
            f"based on the CALL CONTINUITY section in your system prompt."
        )
    elif jd:
        greeting_kickoff = build_vox_greeting_kickoff(name)
    else:
        greeting_kickoff = SARAH_GREETING_KICKOFF

    async def _send_twilio_json(payload: dict) -> None:
        await websocket.send_text(json.dumps(payload))

    async def _on_call_ended(was_dropped: bool, duration_seconds: float) -> None:
        bridge_transcript = getattr(bridge, "transcript", [])
        logger.info("[WS] Call ended dropped=%s duration=%.1fs turns=%d attempt=%d/3",
                    was_dropped, duration_seconds, len(bridge_transcript), retry_num + 1)

        if not was_dropped:
            if duration_seconds < 30:
                was_dropped = True
            else:
                await CallRetryManager.clear(phone)
                return

        if not phone or retry_num >= CallRetryManager.MAX_RETRIES:
            if phone:
                await CallRetryManager.clear(phone)
            return

        new_retry_num = await CallRetryManager.record_drop(
            phone=phone, name=name, jd=jd,
            transcript=bridge_transcript, notes={},
            resume_text=resume_text, recruiter_inputs=recruiter_inputs,
            voice_id=voice_id,
        )
        host_url = settings.PUBLIC_URL.strip()
        from_number = settings.TWILIO_PHONE_NUMBER.strip()
        if not host_url or not from_number:
            return

        delay = RETRY_1_DELAY if new_retry_num == 1 else RETRY_2_DELAY
        state = await CallRetryManager.load(phone)
        CallRetryManager.schedule_callback(
            phone=phone, name=name, jd=jd,
            transcript=state.get("transcript", bridge_transcript),
            notes=state.get("notes", {}),
            resume_text=state.get("resume_text", resume_text),
            recruiter_inputs=state.get("recruiter_inputs", recruiter_inputs),
            voice_id=state.get("voice_id", voice_id),
            retry_num=new_retry_num, delay_seconds=delay,
            host_url=host_url, from_number=from_number,
        )

    bridge = GeminiLiveBridge(
        mode="twilio",
        system_prompt=system_prompt,
        greeting_kickoff=greeting_kickoff,
        on_send_twilio_json=_send_twilio_json,
        on_call_ended=_on_call_ended,
        finalize_consumer=_FinalizeConsumer(),
        candidate_name=name,
        candidate_phone=phone,
        job_description=jd,
        call_sid=call_sid,
        call_channel="twilio",
        interview_context=context,
        resume_text=resume_text,
        voice_id=voice_id,
    )
    bridge_task = asyncio.create_task(bridge.run())

    try:
        async for text in websocket.iter_text():
            try:
                data = json.loads(text)
                bridge.enqueue_twilio_event(data)
            except json.JSONDecodeError:
                pass
    except WebSocketDisconnect:
        pass
    finally:
        _csid = getattr(bridge, "_call_sid", "")
        bridge.close()
        if _csid:
            await cache.set(f"vox:ended:{_csid}", True, timeout=3600)
        if not bridge_task.done():
            try:
                await asyncio.wait_for(bridge_task, timeout=30.0)
            except asyncio.TimeoutError:
                bridge_task.cancel()
                try:
                    await bridge_task
                except Exception:
                    pass
            except Exception:
                pass


app.add_api_websocket_route("/media-stream/{token}", _twilio_websocket_handler)
app.add_api_websocket_route("/ws/media-stream/{token}", _twilio_websocket_handler)


# ---------------------------------------------------------------------------
# Campaign background caller
# ---------------------------------------------------------------------------

async def _run_campaign_caller(campaign_id: int) -> None:
    logger.info("[Campaign-%d] Caller task started", campaign_id)
    try:
        host_url = settings.PUBLIC_URL.strip()
        from_number = settings.TWILIO_PHONE_NUMBER.strip()

        if not host_url or not from_number:
            logger.error("[Campaign-%d] PUBLIC_URL or TWILIO_PHONE_NUMBER not configured", campaign_id)
            async with AsyncSessionLocal() as db:
                await db.execute(update(Campaign).where(Campaign.id == campaign_id).values(status=Campaign.PAUSED))
                await db.commit()
            return

        while True:
            async with AsyncSessionLocal() as db:
                res = await db.execute(select(Campaign).where(Campaign.id == campaign_id))
                campaign = res.scalars().first()
            if not campaign or campaign.status != Campaign.RUNNING:
                logger.info("[Campaign-%d] Status=%s — stopping", campaign_id, getattr(campaign, "status", "?"))
                break

            async with AsyncSessionLocal() as db:
                res = await db.execute(
                    select(CampaignCandidate)
                    .where(CampaignCandidate.campaign_id == campaign_id,
                           CampaignCandidate.is_valid == True,
                           CampaignCandidate.is_duplicate == False,
                           CampaignCandidate.status == CampaignCandidate.PENDING)
                    .order_by(CampaignCandidate.id)
                    .limit(1)
                )
                candidate = res.scalars().first()

            if candidate is None:
                logger.info("[Campaign-%d] All candidates processed — marking completed", campaign_id)
                async with AsyncSessionLocal() as db:
                    await db.execute(
                        update(Campaign).where(Campaign.id == campaign_id)
                        .values(status=Campaign.COMPLETED, completed_at=_now())
                    )
                    await db.commit()
                break

            logger.info("[Campaign-%d] Next: %s (%s)", campaign_id, candidate.name, candidate.phone)

            async with AsyncSessionLocal() as db:
                await db.execute(
                    update(CampaignCandidate).where(CampaignCandidate.id == candidate.id)
                    .values(status=CampaignCandidate.CALLING, called_at=_now())
                )
                await db.commit()

            call_sid = ""
            cand_id = candidate.id
            cand_name = candidate.name
            cand_phone = candidate.phone
            try:
                result = await _place_call(
                    to_number=cand_phone, from_number=from_number, host_url=host_url,
                    jd=campaign.job_description or "General screening call for open positions at our company.",
                    name=cand_name, voice_id=campaign.voice_id,
                )
                call_sid = result.get("call_sid") or ""
                if not call_sid:
                    raise RuntimeError("_place_call returned no call_sid")

                async with AsyncSessionLocal() as db:
                    await db.execute(
                        update(CampaignCandidate).where(CampaignCandidate.id == cand_id)
                        .values(call_sid=call_sid)
                    )
                    await db.commit()

                await cache.set(
                    f"vox:campaign_call:{call_sid}",
                    {"campaign_id": campaign_id, "candidate_id": cand_id},
                    timeout=3600,
                )
                logger.info("[Campaign-%d] Placed call to %s → SID %s", campaign_id, cand_name, call_sid)

                await _wait_for_call_end(call_sid, timeout=600)
                logger.info("[Campaign-%d] Call ended for %s (SID %s)", campaign_id, cand_name, call_sid)
                await _sync_call_result(cand_id, call_sid)

            except asyncio.CancelledError:
                raise
            except Exception as exc:
                logger.error("[Campaign-%d] Call failed for %s: %s", campaign_id, cand_name, exc, exc_info=True)
                async with AsyncSessionLocal() as db:
                    await db.execute(
                        update(CampaignCandidate).where(CampaignCandidate.id == cand_id)
                        .values(status=CampaignCandidate.FAILED)
                    )
                    await db.commit()

            delay = campaign.delay_seconds
            logger.info("[Campaign-%d] Sleeping %ds before next", campaign_id, delay)
            await asyncio.sleep(delay)

    except asyncio.CancelledError:
        logger.info("[Campaign-%d] Caller task cancelled", campaign_id)
    except Exception as exc:
        logger.error("[Campaign-%d] Caller crashed: %s", campaign_id, exc, exc_info=True)


async def _wait_for_call_end(call_sid: str, timeout: int = 480) -> None:
    elapsed = 0
    poll = 3
    while elapsed < timeout:
        await asyncio.sleep(poll)
        elapsed += poll
        async with AsyncSessionLocal() as db:
            res = await db.execute(
                select(CallSession.ended_at).where(CallSession.call_sid == call_sid)
            )
            ended_at = res.scalar()
            if ended_at is not None:
                logger.info("[Campaign] Call %s ended after ~%ds", call_sid, elapsed)
                return
            res2 = await db.execute(
                select(CampaignCandidate.status).where(CampaignCandidate.call_sid == call_sid)
            )
            status = res2.scalar()
            if status and status != CampaignCandidate.CALLING:
                logger.info("[Campaign] Call %s ended (candidate status) after ~%ds", call_sid, elapsed)
                return
    logger.warning("[Campaign] Call %s timed out after %ds", call_sid, timeout)


async def resume_running_campaigns() -> None:
    try:
        async with AsyncSessionLocal() as db:
            res = await db.execute(select(Campaign.id).where(Campaign.status == Campaign.RUNNING))
            running_ids = res.scalars().all()
        for cid in running_ids:
            existing = _CAMPAIGN_TASKS.get(cid)
            if existing and not existing.done():
                continue
            async with AsyncSessionLocal() as db:
                await db.execute(
                    update(CampaignCandidate)
                    .where(CampaignCandidate.campaign_id == cid,
                           CampaignCandidate.status == CampaignCandidate.CALLING)
                    .values(status=CampaignCandidate.PENDING)
                )
                await db.commit()
            logger.info("[Startup] Resuming campaign %d", cid)
            task = asyncio.create_task(_run_campaign_caller(cid))
            _CAMPAIGN_TASKS[cid] = task
    except Exception as exc:
        logger.error("[Startup] Failed to resume running campaigns: %s", exc)


# ---------------------------------------------------------------------------
# Candidate validation helpers (campaign upload)
# ---------------------------------------------------------------------------

_E164_NORM_RE = re.compile(r"[\s\-\(\)\.\+]")


def _normalize_phone(raw: str) -> str:
    phone = _E164_NORM_RE.sub("", str(raw or "").strip())
    if not phone:
        return raw
    if re.match(r"^\+?91[6-9]\d{9}$", phone):
        return f"+91{phone[-10:]}"
    if re.match(r"^[6-9]\d{9}$", phone):
        return f"+91{phone}"
    if re.match(r"^\d{10}$", phone):
        return f"+1{phone}"
    if re.match(r"^\+?1\d{10}$", phone):
        return f"+1{phone[-10:]}"
    if raw.strip().startswith("+") and re.match(r"^\d{7,14}$", phone):
        return f"+{phone}"
    return raw


def _validate_candidates(rows: list[dict]) -> dict:
    valid, invalid, duplicates = [], [], []
    seen: set[str] = set()
    for row in rows:
        name = str(row.get("name") or "").strip()
        raw_phone = str(row.get("phone") or "").strip()
        if not name and not raw_phone:
            continue
        phone = _normalize_phone(raw_phone)
        error = None
        if not name:
            error = "Missing name"
        elif not raw_phone:
            error = "Missing phone number"
        elif not _E164_RE.match(phone):
            error = f"Invalid phone: {raw_phone}"
        entry = {"name": name, "phone": phone, "raw_phone": raw_phone}
        if error:
            invalid.append({**entry, "error": error})
        elif phone in seen:
            duplicates.append({**entry, "error": "Duplicate phone number"})
        else:
            seen.add(phone)
            valid.append(entry)
    return {"valid": valid, "invalid": invalid, "duplicates": duplicates}


def _calc_stats(cands: list) -> dict:
    total = len(cands)
    completed = sum(1 for c in cands if c.status == CampaignCandidate.COMPLETED)
    return {
        "total": total,
        "pending": sum(1 for c in cands if c.status == CampaignCandidate.PENDING),
        "calling": sum(1 for c in cands if c.status == CampaignCandidate.CALLING),
        "completed": completed,
        "failed": sum(1 for c in cands if c.status == CampaignCandidate.FAILED),
        "interested": sum(1 for c in cands if c.interest_level == "INTERESTED"),
        "not_interested": sum(1 for c in cands if c.interest_level == "NOT_INTERESTED"),
        "completion_pct": round(completed / total * 100) if total else 0,
    }
